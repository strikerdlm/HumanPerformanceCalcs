# -*- coding: utf-8 -*-
"""
Created on Mon Apr  3 12:53:55 2023

@author: User
"""

import openpyxl
import csv

def calcular_mssq_short(seccion_A, seccion_B):
    MSA = sum(seccion_A) * 9 / (9 - seccion_A.count(0))
    MSB = sum(seccion_B) * 9 / (9 - seccion_B.count(0))
    MSSQ_short_raw_score = MSA + MSB
    return MSA, MSB, MSSQ_short_raw_score

def convertir_a_enteros(valores):
    conversion_key = {
        'No aplica': 0,
        'Nunca': 0,
        'Casi Nunca': 1,
        'A veces': 2,
        'Con frecuencia': 3
    }
    enteros = [conversion_key.get(valor, valor) for valor in valores]
    return enteros

def leer_datos_xlsx(nombre_archivo):
    resultados = []
    wb = openpyxl.load_workbook(nombre_archivo)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    for row in ws.iter_rows(min_row=2):
        row_dict = {headers[i]: cell.value for i, cell in enumerate(row)}
        seccion_A = convertir_a_enteros([row_dict['Automoviles'], row_dict['Buses o microbuses'], row_dict['Trenes'], row_dict['Aeronaves'], row_dict['Botes Pequeños'], row_dict['Embarcaciones'], row_dict['Columpios'], row_dict['Juegos Infantiles de plaza'], row_dict['Toboganes, juegos mecánicos de  parques de diversiones']])
        seccion_B = convertir_a_enteros([row_dict['Automóviles'], row_dict['Buses o microbuses2'], row_dict['Trenes2'], row_dict['Aeronaves2'], row_dict['Botes pequeños2'], row_dict['Embarcaciones2'], row_dict['Columpios2'], row_dict['Juegos infantiles de plaza2'], row_dict['Toboganes, juegos mecánicos de  parques de diversiones2']])
        MSA, MSB, mssq_short_raw_score = calcular_mssq_short(seccion_A, seccion_B)
        resultados.append({'ID': row_dict['ID'], 'MSA': MSA, 'MSB': MSB, 'MSSQ-short raw score': mssq_short_raw_score})

    return resultados

def escribir_resultados_csv(resultados, nombre_archivo_salida):
    with open(nombre_archivo_salida, 'w', newline='', encoding='utf-8') as csvfile:
        fieldnames = ['ID', 'MSA', 'MSB', 'MSSQ-short raw score']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        for resultado in resultados:
            writer.writerow(resultado)

nombre_archivo = r"C:\Users\User\Downloads\CUESTIONARIO ABREVIADO DE SUSCEPTIBILIDAD A LA CINETOSIS ADAPTADO AL ESPAÑOL(1-1).xlsx"
resultados = leer_datos_xlsx(nombre_archivo)
nombre_archivo_salida = "resultados_mssq.csv"
escribir_resultados_csv(resultados, nombre_archivo_salida)
