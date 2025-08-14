# -*- coding: utf-8 -*-
"""
Motion Sickness Susceptibility Questionnaire (MSSQ) Calculator
Created on: Original date unknown
Updated for portability and usability

@author: Diego Malpica

Processes MSSQ questionnaire data from Excel files and calculates
motion sickness susceptibility scores.
Configurable file paths for improved portability.
"""

import csv
import openpyxl
import os
import argparse
from pathlib import Path
from typing import List, Dict, Any, Optional


def convertir_a_enteros(lista: List[Any]) -> List[int]:
    """
    Convert a list of values to integers, handling None values.
    
    Args:
        lista (List[Any]): List of values to convert
        
    Returns:
        List[int]: List of integers with None values converted to 0
    """
    return [int(valor) if valor is not None else 0 for valor in lista]


def calcular_mssq_short(seccion_A: List[int], seccion_B: List[int]) -> tuple[float, float, float]:
    """
    Calculate MSSQ-short scores for sections A and B.
    
    Args:
        seccion_A (List[int]): Section A responses
        seccion_B (List[int]): Section B responses
        
    Returns:
        tuple: (MSA score, MSB score, total MSSQ-short raw score)
    """
    # Calculate MSA (Motion Sickness A - childhood experiences)
    suma_A = sum(seccion_A)
    num_items_A = len([x for x in seccion_A if x > 0])
    MSA = suma_A / max(num_items_A, 1)  # Avoid division by zero
    
    # Calculate MSB (Motion Sickness B - adult experiences)
    suma_B = sum(seccion_B)
    num_items_B = len([x for x in seccion_B if x > 0])
    MSB = suma_B / max(num_items_B, 1)  # Avoid division by zero
    
    # Calculate total MSSQ-short raw score
    mssq_short_raw_score = (suma_A + suma_B) / max((num_items_A + num_items_B), 1)
    
    return MSA, MSB, mssq_short_raw_score


def leer_datos_xlsx(file_path: str) -> List[Dict[str, Any]]:
    """
    Read MSSQ data from an Excel file and calculate scores.
    
    Args:
        file_path (str): Path to the Excel file
        
    Returns:
        List[Dict[str, Any]]: List of results with calculated scores
        
    Raises:
        FileNotFoundError: If the file doesn't exist
        ValueError: If required columns are missing
    """
    if not Path(file_path).exists():
        raise FileNotFoundError(f"Data file not found: {file_path}")
    
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
    except Exception as e:
        raise ValueError(f"Error reading Excel file: {e}")
    
    # Get headers from first row
    headers = [cell.value for cell in ws[1]]
    
    # Expected column names for section A (childhood experiences)
    section_a_columns = [
        'Automoviles', 'Buses o microbuses', 'Trenes', 'Aeronaves', 
        'Botes Pequeños', 'Embarcaciones', 'Columpios', 
        'Juegos Infantiles de plaza', 'Toboganes, juegos mecánicos de  parques de diversiones'
    ]
    
    # Expected column names for section B (adult experiences)
    section_b_columns = [
        'Automóviles', 'Buses o microbuses2', 'Trenes2', 'Aeronaves2', 
        'Botes pequeños2', 'Embarcaciones2', 'Columpios2', 
        'Juegos infantiles de plaza2', 'Toboganes, juegos mecánicos de  parques de diversiones2'
    ]
    
    # Check if required columns exist
    required_columns = ['ID'] + section_a_columns + section_b_columns
    missing_columns = [col for col in required_columns if col not in headers]
    if missing_columns:
        print(f"Warning: Missing columns: {missing_columns}")
        print(f"Available columns: {headers}")
    
    resultados = []
    
    for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
        try:
            # Create dictionary from row data
            row_dict = {headers[i]: cell.value for i, cell in enumerate(row) if i < len(headers)}
            
            # Extract section A data (with error handling for missing columns)
            seccion_A = []
            for col in section_a_columns:
                value = row_dict.get(col, 0)
                seccion_A.append(value)
            seccion_A = convertir_a_enteros(seccion_A)
            
            # Extract section B data (with error handling for missing columns)
            seccion_B = []
            for col in section_b_columns:
                value = row_dict.get(col, 0)
                seccion_B.append(value)
            seccion_B = convertir_a_enteros(seccion_B)
            
            # Calculate MSSQ scores
            MSA, MSB, mssq_short_raw_score = calcular_mssq_short(seccion_A, seccion_B)
            
            # Get ID (with fallback if missing)
            participant_id = row_dict.get('ID', f'Participant_{row_num}')
            
            resultados.append({
                'ID': participant_id,
                'MSA': MSA,
                'MSB': MSB,
                'MSSQ-short raw score': mssq_short_raw_score,
                'Section_A_responses': seccion_A,
                'Section_B_responses': seccion_B
            })
            
        except Exception as e:
            print(f"Warning: Error processing row {row_num}: {e}")
            continue
    
    print(f"Processed {len(resultados)} records successfully")
    return resultados


def escribir_resultados_csv(resultados: List[Dict[str, Any]], output_file: str, 
                           include_responses: bool = False) -> None:
    """
    Write MSSQ results to a CSV file.
    
    Args:
        resultados (List[Dict[str, Any]]): Results to write
        output_file (str): Output CSV file path
        include_responses (bool): Whether to include individual responses
    """
    if not resultados:
        print("No results to write")
        return
    
    # Create output directory if it doesn't exist
    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Define fieldnames
    fieldnames = ['ID', 'MSA', 'MSB', 'MSSQ-short raw score']
    if include_responses:
        fieldnames.extend(['Section_A_responses', 'Section_B_responses'])
    
    try:
        with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            
            for resultado in resultados:
                # Create row data
                row_data = {
                    'ID': resultado['ID'],
                    'MSA': f"{resultado['MSA']:.4f}",
                    'MSB': f"{resultado['MSB']:.4f}",
                    'MSSQ-short raw score': f"{resultado['MSSQ-short raw score']:.4f}"
                }
                
                if include_responses:
                    row_data['Section_A_responses'] = str(resultado['Section_A_responses'])
                    row_data['Section_B_responses'] = str(resultado['Section_B_responses'])
                
                writer.writerow(row_data)
        
        print(f"Results written to: {output_file}")
        
    except Exception as e:
        raise ValueError(f"Error writing CSV file: {e}")


def generate_summary_statistics(resultados: List[Dict[str, Any]]) -> Dict[str, float]:
    """
    Generate summary statistics for MSSQ scores.
    
    Args:
        resultados (List[Dict[str, Any]]): Results data
        
    Returns:
        Dict[str, float]: Summary statistics
    """
    if not resultados:
        return {}
    
    msa_scores = [r['MSA'] for r in resultados]
    msb_scores = [r['MSB'] for r in resultados]
    mssq_scores = [r['MSSQ-short raw score'] for r in resultados]
    
    def calculate_stats(scores: List[float]) -> Dict[str, float]:
        return {
            'mean': sum(scores) / len(scores),
            'min': min(scores),
            'max': max(scores),
            'count': len(scores)
        }
    
    return {
        'MSA': calculate_stats(msa_scores),
        'MSB': calculate_stats(msb_scores),
        'MSSQ_total': calculate_stats(mssq_scores)
    }


def main():
    """Main function to run the MSSQ calculator."""
    parser = argparse.ArgumentParser(description="MSSQ Calculator - Process motion sickness questionnaire data")
    parser.add_argument("--input_file", type=str,
                       default=os.getenv("MSSQ_INPUT_FILE", "data/mssq_questionnaire.xlsx"),
                       help="Path to the input Excel file")
    parser.add_argument("--output_file", type=str,
                       default=os.getenv("MSSQ_OUTPUT_FILE", "output/resultados_mssq.csv"),
                       help="Path to the output CSV file")
    parser.add_argument("--include_responses", action="store_true",
                       help="Include individual responses in output")
    parser.add_argument("--summary", action="store_true",
                       help="Generate summary statistics")
    
    args = parser.parse_args()
    
    try:
        print("MSSQ Calculator")
        print("=" * 20)
        print(f"Input file: {args.input_file}")
        print(f"Output file: {args.output_file}")
        print()
        
        # Process the Excel file
        print("Reading and processing MSSQ data...")
        resultados = leer_datos_xlsx(args.input_file)
        
        if not resultados:
            print("No valid data found in the input file")
            return 1
        
        # Write results to CSV
        print("Writing results...")
        escribir_resultados_csv(resultados, args.output_file, args.include_responses)
        
        # Generate summary statistics if requested
        if args.summary:
            print("\nGenerating summary statistics...")
            stats = generate_summary_statistics(resultados)
            
            print("\nSummary Statistics:")
            print("-" * 30)
            for category, category_stats in stats.items():
                print(f"\n{category}:")
                for stat_name, value in category_stats.items():
                    if stat_name == 'count':
                        print(f"  {stat_name}: {int(value)}")
                    else:
                        print(f"  {stat_name}: {value:.4f}")
            
            # Save summary to file
            summary_file = Path(args.output_file).parent / "mssq_summary.txt"
            with open(summary_file, 'w') as f:
                f.write("MSSQ Summary Statistics\n")
                f.write("=" * 25 + "\n\n")
                for category, category_stats in stats.items():
                    f.write(f"{category}:\n")
                    for stat_name, value in category_stats.items():
                        if stat_name == 'count':
                            f.write(f"  {stat_name}: {int(value)}\n")
                        else:
                            f.write(f"  {stat_name}: {value:.4f}\n")
                    f.write("\n")
            
            print(f"Summary statistics saved to: {summary_file}")
        
        print(f"\nProcessing completed successfully!")
        print(f"Processed {len(resultados)} participants")
        
    except Exception as e:
        print(f"Error: {e}")
        return 1
    
    return 0


if __name__ == "__main__":
    exit(main())
